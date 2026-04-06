import openpyxl, json, sys

wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
blocks = []

for name in wb.sheetnames:
    ws = wb[name]
    dates = []
    for col in range(2, 9):
        cell = ws.cell(row=3, column=col).value
        if cell:
            dates.append(cell.strftime('%Y-%m-%d'))
        else:
            dates.append(None)

    for row_idx in range(4, 19):
        hour_label = ws.cell(row=row_idx, column=1).value
        if not hour_label:
            continue
        hour = int(hour_label.split(' - ')[0])

        for col_idx in range(2, 9):
            category = ws.cell(row=row_idx, column=col_idx).value
            day_idx = col_idx - 2
            if category and isinstance(category, str) and dates[day_idx]:
                cat = category.strip()
                is_prep = False
                if cat.lower().startswith('prep '):
                    is_prep = True
                    cat = cat[5:]  # Remove "Prep " prefix
                    # Capitalize first letter to match category names
                    cat = cat[0].upper() + cat[1:] if cat else cat
                blocks.append({
                    'date': dates[day_idx],
                    'hour': hour,
                    'category': cat,
                    'is_prep': is_prep
                })

print(json.dumps(blocks))
